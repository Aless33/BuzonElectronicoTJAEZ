"""
Generación de reportes en Excel para el Buzón Electrónico TJAEZ.
"""

import datetime

from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from web.models import Etiqueta

ENCABEZADOS = [
    "Número de sobre",
    "UUID",
    "Dígito verificador",
    "Fecha y hora de depósito",
]


def generar_workbook_depositos(fecha):
    """
    Construye un Workbook de openpyxl con las etiquetas depositadas
    en la fecha dada (objeto datetime.date).
    """
    etiquetas = (
        Etiqueta.objects.filter(
            estado=Etiqueta.ESTADO_DEPOSITADO,
            fecha_deposito__date=fecha,
        )
        .order_by("fecha_deposito")
    )

    wb = Workbook()
    ws = wb.active
    ws.title = f"Depositos {fecha.isoformat()}"

    ws.append(ENCABEZADOS)
    for celda in ws[1]:
        celda.font = Font(bold=True)

    for etiqueta in etiquetas:
        fecha_local = timezone.localtime(etiqueta.fecha_deposito)
        ws.append([
            etiqueta.numero_sobre,
            str(etiqueta.uuid),
            etiqueta.digito_verificador,
            fecha_local.strftime("%Y-%m-%d %H:%M:%S"),
        ])

    # Autoajuste simple de ancho de columnas
    for i, encabezado in enumerate(ENCABEZADOS, start=1):
        columna = get_column_letter(i)
        ws.column_dimensions[columna].width = max(len(encabezado) + 4, 18)

    # Fila de resumen al final
    ws.append([])
    ws.append(["Total de sobres depositados:", etiquetas.count()])
    ws.cell(
        row=ws.max_row, column=1
    ).font = Font(bold=True)

    return wb